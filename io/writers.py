"""
Output writers: tab-delimited TXT and TMX 1.4.

TMX structure
-------------
  <tmx version="1.4">
    <header .../>
    <body>
      <tu tuid="N">
        <tuv xml:lang="en-GB"><seg>source</seg></tuv>
        <tuv xml:lang="pl-PL"><seg>target</seg></tuv>
      </tu>
    </body>
  </tmx>
"""

from __future__ import annotations

import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from ..utils import tmx_timestamp


# ---------------------------------------------------------------------------
# Tab-delimited text
# ---------------------------------------------------------------------------

def write_txt(pairs: list[tuple[str, ...]],
              path: str | Path,
              bom: bool = True,
              source_note: str = "") -> None:
    """
    Write aligned pairs to a tab-delimited UTF-8 text file.

    Each element of *pairs* is a tuple of strings (source, target[, extra…]).
    An optional *source_note* is appended as the last column on every row,
    matching the Perl behaviour of recording the source filename.
    """
    encoding = "utf-8-sig" if bom else "utf-8"
    with open(path, "w", encoding=encoding, newline="\n") as f:
        for pair in pairs:
            row = list(pair)
            if source_note:
                row.append(source_note)
            f.write("\t".join(row) + "\n")


# ---------------------------------------------------------------------------
# TMX 1.4
# ---------------------------------------------------------------------------

def write_tmx(pairs: list[tuple[str, ...]],
              path: str | Path,
              langs: list[str],
              creator_id: str = "",
              creation_date: Optional[str] = None,
              note: str = "",
              skip_half_empty: bool = True) -> None:
    """
    Write aligned pairs to a TMX 1.4 file.

    Parameters
    ----------
    pairs:
        Each tuple contains one text per language, in the same order as *langs*.
    langs:
        BCP-47 / TMX language codes, e.g. ["en-GB", "pl-PL"].
    creator_id:
        Optional creator identifier written into the TMX header.
    creation_date:
        TMX-format timestamp (yyyymmddThhmmssZ).  Defaults to now.
    note:
        Optional note added as a <prop type="Txt::Note"> to every TU.
    skip_half_empty:
        If True, TUs where any language segment is empty are omitted.
    """
    if creation_date is None:
        creation_date = tmx_timestamp()

    tmx = ET.Element("tmx", version="1.4")
    ET.SubElement(
        tmx, "header",
        creationtool="lfp_aligner",
        creationtoolversion="1.0",
        datatype="PlainText",
        segtype="sentence",
        adminlang="en-US",
        srclang=langs[0] if langs else "en",
        creationdate=creation_date,
        creationid=creator_id,
        **{"o-tmf": "TMX"},
    )
    body = ET.SubElement(tmx, "body")

    tuid = 0
    for pair in pairs:
        texts = list(pair)

        if skip_half_empty and any(not t.strip() for t in texts):
            continue

        tuid += 1
        tu = ET.SubElement(body, "tu", tuid=str(tuid))

        if note:
            prop = ET.SubElement(tu, "prop")
            prop.set("type", "Txt::Note")
            prop.text = note

        for lang, text in zip(langs, texts):
            tuv = ET.SubElement(tu, "tuv")
            tuv.set("{http://www.w3.org/XML/1998/namespace}lang", lang)
            seg = ET.SubElement(tuv, "seg")
            seg.text = text.strip()

    # Pretty-print (Python 3.9+); fall back gracefully on older versions
    try:
        ET.indent(tmx, space="  ")
    except AttributeError:
        pass

    tree = ET.ElementTree(tmx)
    tree.write(str(path), encoding="utf-8", xml_declaration=True)


# ---------------------------------------------------------------------------
# XLSX (for manual review)
# ---------------------------------------------------------------------------

def write_xlsx(pairs: list[tuple[str, ...]],
               path: str | Path,
               langs: list[str],
               max_rows: int = 1_048_576) -> None:
    """
    Write aligned pairs to an .xlsx spreadsheet for manual review.

    Requires openpyxl.  Columns are sized per language count.
    Rows beyond *max_rows* are silently truncated.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for XLSX output.  "
                          "Run: pip install openpyxl")

    wb = openpyxl.Workbook()

    # --- Instructions sheet ---
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info["A1"] = (
        "This file was created by LF Aligner (Python port).  "
        "Edit segments in the 'Alignment' sheet, then save.  "
        "Empty rows will be removed on import."
    )
    ws_info["A1"].font = Font(bold=True)

    # --- Alignment sheet ---
    ws = wb.create_sheet("Alignment")

    header_fill = PatternFill("solid", fgColor="DDEEFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    col_width = max(30, 80 // max(len(langs), 1))

    # Header row
    for col_idx, lang in enumerate(langs, start=1):
        cell = ws.cell(row=1, column=col_idx, value=lang)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = col_width

    # Data rows
    for row_idx, pair in enumerate(pairs[:max_rows], start=2):
        for col_idx, text in enumerate(pair, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = wrap

    ws.freeze_panes = "A2"
    wb.save(str(path))
